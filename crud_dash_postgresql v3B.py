import dash
from dash import Dash, dcc, html, dash_table, Input, Output, State
from dash.exceptions import PreventUpdate
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from flask import Flask, json
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import text
from openpyxl import load_workbook
import warnings
from tqdm import tqdm
import io
import base64
import os
import datetime


# app requires "pip install psycopg2" as well

server = Flask(__name__)
app = Dash(__name__, server=server, suppress_callback_exceptions=True)
app.server.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# for your home PostgreSQL test table
app.server.config["SQLALCHEMY_DATABASE_URI"] = "postgresql://postgres:PGPassword001@localhost/crud_dash"

db = SQLAlchemy(app.server)

class Product(db.Model):
    __tablename__ = 'ziltektable'

    MK_Type = db.Column(db.Text, nullable=True, primary_key=True)
    Sheet = db.Column(db.Text, nullable=True,)
    Client = db.Column(db.Text, nullable=True,)
    Country = db.Column(db.Text, nullable=True,)
    Service_date = db.Column(db.Date, nullable=True,)
    Reason_For_Service = db.Column(db.Text, nullable=True,)
    User_ID = db.Column(db.Text, nullable=True,)
    User_Password = db.Column(db.Text, nullable=True,)
    Background_Cap  = db.Column(db.Float, nullable=True,)
    Polystyrene_PS_Cap = db.Column(db.Float, nullable=True,)
    SNR_1142_1042_cm1 = db.Column(db.Float, nullable=True,)
    SNR_2600_2500_cm1 = db.Column(db.Float, nullable=True,)
    Centre_burst_intensity = db.Column(db.Float, nullable=True,)
    Single_beam_spectrum_4200_4500 = db.Column(db.Float, nullable=True,)
    Single_beam_spectrum_2600_3000 = db.Column(db.Float, nullable=True,)


    def __init__(self, MK_Type, Sheet, Client, Country,Service_date,Reason_For_Service,User_ID,User_Password,Background_Cap,Polystyrene_PS_Cap,SNR_1142_1042_cm1,SNR_2600_2500_cm1,Centre_burst_intensity,Single_beam_spectrum_4200_4500,Single_beam_spectrum_2600_3000):
     self.MK_Type = MK_Type
     self.Sheet = Sheet
     self.Client = Client
     self.Country = Country
     self.Service_date = Service_date
     self.Reason_For_Service = Reason_For_Service
     self.User_ID = User_ID
     self.User_Password = User_Password
     self.Background_Cap = Background_Cap
     self.Polystyrene_PS_Cap = Polystyrene_PS_Cap
     self.SNR_1142_1042_cm1 = SNR_1142_1042_cm1
     self.SNR_2600_2500_cm1 = SNR_2600_2500_cm1
     self.Centre_burst_intensity = Centre_burst_intensity
     self.Single_beam_spectrum_4200_4500 = Single_beam_spectrum_4200_4500



# ------------------------------------------------------------------------------------------------

app.layout = html.Div([
    html.Div([
        dcc.Input(
            id='adding-rows-name',
            placeholder='Enter a column name...',
            value='',
            style={'padding': 10}
        ),
        html.Button('Add Column', id='adding-columns-button', n_clicks=0)
    ], style={'height': 50}),

    dcc.Interval(id='interval_pg', interval=86400000*7, n_intervals=0),  # activated once/week or when page refreshed
    html.Div(id='postgres_datatable'),

    html.Button('Add Row', id='editing-rows-button', n_clicks=0),
    html.Button('Save to PostgreSQL', id='save_to_postgres', n_clicks=0),

    # -- from exdash --
    dcc.Upload(
        id='upload-data',
        children=html.Div([
            'Drag and Drop or ',
            html.A('Select Excel File')
        ]),
        style={
            'width': '100%',
            'height': '60px',
            'lineHeight': '60px',
            'borderWidth': '1px',
            'borderStyle': 'dashed',
            'borderRadius': '5px',
            'textAlign': 'center',
            'margin': '10px'
        },
        multiple=False
    ),

    html.Div(id='output-data-upload'),
    dcc.Loading(id="loading-output", type="default", children=[]),
    html.Div(id='table-container'),

    # -- exdash out --

    # Create notification when saving to excel
    html.Div(id='placeholder', children=[]),
    dcc.Store(id="store", data=0),
    dcc.Interval(id='interval', interval=1000),
    dcc.Graph(id='my_graph_year'),
    dcc.Graph(id='my_graph_month'),
    html.Hr()

])

# ------------------------------------------------------------------------------------------------
# from exdashtest

def process_excel(file_path):
    search_strings = [
        'Client',
        'Country',
        'Service date',
        'Reason for Service',
        'RemScan Serial #',
        'User ID',
        'Password',
        'Background Cap (Minimum requirement = 4500 @ Gain = 255)',
        'Polystyrene P/S Cap (Minimum requirement = 4000 @ Gain = 255)',
        'SNR: (1142 - 1042 cm-1) (Recommended requirement = 4500)',
        'SNR: (2600 - 2500 cm-1) ',
        'Centre burst intensity (Interferogram) (Minmum requirement =20,000)'
    ]

    two_cells_away_strings = [
        'Single beam spectrum (Counts: 4200-4500 / Total Counts)x100                  (Minimum requirement = 1%)',
        'Single beam spectrum (Counts: 2600-3000 / Total Counts)x100                  (Minimum requirement = 7%)'
    ]

    df = pd.DataFrame(columns= ['Type', 'Sheet'] +search_strings + two_cells_away_strings)

    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheets = wb.sheetnames

    for sheet_name in sheets:
        sheet = wb[sheet_name]
        found_values = find_values(sheet, search_strings, two_cells_away_strings)

        temp_data = {'Type': '', 'Sheet': ''}

        for string, value in found_values.items():
            if string == 'Service date':
                if value is not None:
                    if isinstance(value, datetime.datetime):
                        value = value.date().strftime('%Y-%m-%d')
                    else:
                        try:
                            value = datetime.datetime.strptime(str(value), '%d/%m/%Y').date().strftime('%Y-%m-%d')
                        except ValueError:
                            value = 'N/A'
            temp_data[string] = value

        temp_df = pd.DataFrame(temp_data, index=[0])

        df = pd.concat([df, temp_df], ignore_index=True)

    df.columns = [
        'MK_Type', 'Sheet', 'Client', 'Country', 'Service_date', 'Reason_for_Service', 'RemScan_Serial', 'User_ID', 'User_Password',
        'Background_Cap', 'Polystyrene_PS_Cap', 'SNR_1142_1042_cm1', 'SNR_2600_2500_cm1', 'Centre_burst_intensity',
        'Single_beam_spectrum_4200_4500', 'Single_beam_spectrum_2600_3000'
    ]

    for col in [
        'Background_Cap', 'Polystyrene_PS_Cap', 'SNR_1142_1042_cm1', 'SNR_2600_2500_cm1', 'Centre_burst_intensity',
        'Single_beam_spectrum_4200_4500', 'Single_beam_spectrum_2600_3000'
    ]:
        df[col] = df[col].apply(lambda x: x if type(x) in [int, float] else None)

    return df

def find_values(sheet, search_strings, two_cells_away_strings):
    found_values = {string: None for string in search_strings + two_cells_away_strings}
    for row in sheet.iter_rows():
        for cell in row:
            for string in search_strings:
                if string in str(cell.value):
                    next_cell_value = sheet.cell(row=cell.row, column=cell.column + 1).value
                    if next_cell_value is not None:
                        found_values[string] = next_cell_value
            for string in two_cells_away_strings:
                if string in str(cell.value):
                    two_cells_away_value = sheet.cell(row=cell.row, column=cell.column + 2).value
                    if two_cells_away_value is not None:
                        found_values[string] = two_cells_away_value
    return found_values

# ------------------------------------------------------------------------------------------------
@app.callback(Output('postgres_datatable', 'children'),
              [Input('interval_pg', 'n_intervals')])
def populate_datatable(n_intervals):
    df = pd.read_sql_table('ziltektable', con=db.engine)
    df['Service_date'] = pd.to_datetime(df['Service_date']).dt.strftime("%Y-%m-%d")
    return [
        dash_table.DataTable(
            id='our-table',
            columns=[{
                         'name': str(x),
                         'id': str(x),
                         'deletable': False,
            }
                     for x in df.columns],
            data=df.to_dict('records'),
            editable=True,
            row_deletable=True,
            filter_action="native",
            sort_action="native",  # give user capability to sort columns
            sort_mode="single",  # sort across 'multi' or 'single' columns
            page_action='none',  # render all of the data at once. No paging.
            style_table={'height': '450px', 'overflowY': 'auto'},
            style_cell={'textAlign': 'left', 'minWidth': '170px', 'width': '100px', 'maxWidth': '100px'},
        ),
    ]


@app.callback(
    Output('our-table', 'columns'),
    [Input('adding-columns-button', 'n_clicks')],
    [State('adding-rows-name', 'value'),
     State('our-table', 'columns')],
    prevent_initial_call=True)
def add_columns(n_clicks, value, existing_columns):
    if n_clicks > 0:
        existing_columns.append({
            'name': value, 'id': value,
            'renamable': True, 'deletable': True
        })
    return existing_columns


@app.callback(
    [Output('our-table', 'data'),
     Output('output-data-upload', 'children'),
     Output('loading-output', 'children')],
    [Input('editing-rows-button', 'n_clicks'),
     Input('upload-data', 'contents')],
    [State('our-table', 'data'),
     State('our-table', 'columns'),
     State('upload-data', 'filename'),
     State('upload-data', 'last_modified')],
    prevent_initial_call=True)
def add_row(n_clicks, contents, rows, columns, filename, last_modified):
    if n_clicks > 0:
        rows.append({c['id']: '' for c in columns})
        
        return rows, None, None
    
    else:
        try:
            if contents is None:
                raise PreventUpdate
            
            content_decoded = base64.b64decode(contents.split(",")[1])
            file_buffer = io.BytesIO(content_decoded)
            df = process_excel(file_buffer)
            
            table = dcc.Loading(id="table-loading", type="default", children=[
                dash_table.DataTable(
                    id='datatable',
                    columns=[{
                                'name': str(col), 
                                'id': str(col),
                                'deletable': False,
                                } for col in df.columns],
                    data=df.to_dict('records'),
                    page_size=10  # Adjust as needed
                )
            ])

            # Convert the rows object to a DataFrame before concatenating it to the df DataFrame.
            rows_df = pd.DataFrame(rows)

            # Concatenate the DataFrames using the concat() method.
            rows = pd.concat([rows_df, df], ignore_index=True)
            rows_json = rows.to_json(orient='records')
            rows_json = json.loads(rows_json)

            return rows_json, html.Div([
                    html.H4(f'File Name: {filename}'),
                    html.P('Data cleaning and extraction completed. Rows were added to the table')
                ]), None

        except Exception as e:
            print(f"Error in callback: {str(e)}")
            return [
                rows,
                html.Div([
                    html.P('An error occurred while processing the file. Please check the file format and try again.'),
                ]),
                None
            ]

# # -- from exdash -- -

# @app.callback(
#     [Output('output-data-upload', 'children'),
#     Output('loading-output', 'children'),
#     Output('table-container', 'children')],  # New output for the table
#     Input('upload-data', 'contents'),
#     [State('upload-data', 'filename'),
#     State('upload-data', 'last_modified')]
# )
# def update_output(contents, filename, last_modified):
#     try:
#         if contents is None:
#             raise PreventUpdate

#         # directory = 'uploads'
#         # os.makedirs(directory, exist_ok=True)
#         # file_path = os.path.join(directory, filename)

#         # Decode the base64 encoded content
#         content_decoded = base64.b64decode(contents.split(",")[1])
#         file_buffer = io.BytesIO(content_decoded)
#         df = process_excel(file_buffer)

#         # with open(file_path, 'wb') as f:
#         #     f.write(content_decoded)

#         # df = process_excel(file_path)
#         # cleaned_csv_path = file_path.replace('.xlsx', '_cleaned.csv')
#         # df.to_csv(cleaned_csv_path, index=False)

#         # Create the Dash table component with the cleaned CSV data
#         table = dcc.Loading(id="table-loading", type="default", children=[
#             dash_table.DataTable(
#                 id='datatable',
#                 columns=[{'name': col, 'id': col} for col in df.columns],
#                 data=df.to_dict('records'),
#                 page_size=10  # Adjust as needed
#             )
#         ])

#         return [
#             html.Div([
#                 html.H4(f'File Name: {filename}'),
#                 html.P('Data cleaning and extraction completed.'),
#                 # html.A('Download Cleaned CSV', href=f'/{cleaned_csv_path}'),
#             ]),
#             None,
#             table  # Return the table as an output
#         ]
#     except Exception as e:
#         print(f"Error in callback: {str(e)}")
#         return [
#             html.Div([
#                 html.P('An error occurred while processing the file. Please check the file format and try again.'),
#             ]),
#             None,
#             None  # Return None for the table in case of an error
#         ]
    
# # -- exdash out --

@app.callback(
    Output('my_graph_year', 'figure'),
    [Input('our-table', 'data')])
def display_graph_year(data):
    if data:
        df_fig = pd.DataFrame(data)
        if 'MK_Type' in df_fig.columns and 'Service_date' in df_fig.columns:
            df_fig['Year'] = pd.to_datetime(df_fig['Service_date']).dt.year
            yearly_counts = df_fig.groupby('Year')['MK_Type'].count().reset_index()
            fig = go.Figure(data=[
                go.Bar(x=yearly_counts['Year'], y=yearly_counts['MK_Type'])
            ])
            fig.update_layout(
                title='Number of Tests Per Year',
                xaxis_title='Year',
                yaxis_title='Number of Tests',
                showlegend=False
            )
            return fig

    # If data is missing or columns don't exist, return an empty figure
    return {'data': []}

@app.callback(
    Output('my_graph_month', 'figure'),
    [Input('my_graph_year', 'clickData'), Input('our-table', 'data')],
)
def display_graph_month(clickData, data):
    if clickData is None:
        return {'data': []}

    selected_year = clickData['points'][0]['x']

    if data:
        df_fig = pd.DataFrame(data)
        if 'MK_Type' in df_fig.columns and 'Service_date' in df_fig.columns:
            df_fig['Year'] = pd.to_datetime(df_fig['Service_date']).dt.year
            df_fig['Month'] = pd.to_datetime(df_fig['Service_date']).dt.month
            filtered_data = df_fig[df_fig['Year'] == selected_year]
            monthly_counts = filtered_data.groupby('Month')['MK_Type'].count().reset_index()

            # Define a dictionary to map month numbers to month names
            month_names = {
                1: 'January',
                2: 'February',
                3: 'March',
                4: 'April',
                5: 'May',
                6: 'June',
                7: 'July',
                8: 'August',
                9: 'September',
                10: 'October',
                11: 'November',
                12: 'December',
            }

            # Map month numbers to month names
            monthly_counts['Month'] = monthly_counts['Month'].map(month_names)

            fig = go.Figure(data=[
                go.Bar(x=monthly_counts['Month'], y=monthly_counts['MK_Type'])
            ])
            fig.update_layout(
                title=f'Number of Tests Per Month in {selected_year}',
                xaxis_title='Month',
                yaxis_title='Number of Tests',
                showlegend=False
            )
            return fig

    # If data is missing or columns don't exist, return an empty figure
    return {'data': []}



@app.callback(
    [Output('placeholder', 'children'),
     Output("store", "data")],
    [Input('save_to_postgres', 'n_clicks'),
     Input("interval", "n_intervals")],
    [State('our-table', 'data'),
     State('store', 'data')],
    prevent_initial_call=True)
def df_to_csv(n_clicks, n_intervals, dataset, s):
    output = html.Plaintext("The data has been saved to your PostgreSQL database.",
                            style={'color': 'green', 'font-weight': 'bold', 'font-size': 'large'})
    no_output = html.Plaintext("", style={'margin': "0px"})

    input_triggered = dash.callback_context.triggered[0]["prop_id"].split(".")[0]

    if input_triggered == "save_to_postgres":
        s = 6
        pg = pd.DataFrame(dataset)
        
        # Ensure the 'Service_date' column is in the proper date format
        pg['Service_date'] = pd.to_datetime(pg['Service_date'], format="%Y-%m-%d")
        
        # Use the SQLAlchemy ORM to insert the data into PostgreSQL
        pg.to_sql("ziltektable", con=db.engine, if_exists='replace', index=False)
        
        # Alter the 'Service_date' column to change its datatype to DATE
        alter_sql = text('ALTER TABLE ziltektable ALTER COLUMN "Service_date" TYPE DATE')
        db.session.execute(alter_sql)
        db.session.commit()  # Commit the transaction
        
        return output, s
    elif input_triggered == 'interval' and s > 0:
        s = s - 1
        if s > 0:
            return output, s
        else:
            return no_output, s
    elif s == 0:
        return no_output, s
    
if __name__ == '__main__':
    app.run_server(debug=True)