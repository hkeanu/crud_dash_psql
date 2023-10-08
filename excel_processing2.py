import os
import warnings
import pandas as pd
from openpyxl import load_workbook
import datetime
from tqdm import tqdm
import time

warnings.simplefilter("ignore", category=UserWarning)


class ExcelFileCombiner:
    def __init__(self, file_paths, file_types, output_file, search_strings, two_cells_away_strings):
        self.file_paths = file_paths
        self.file_types = file_types
        self.output_file = output_file
        self.search_strings = search_strings
        self.two_cells_away_strings = two_cells_away_strings
        
    def find_values(self, sheet):
        # Find values in the sheet based on search strings
        found_values = {string: None for string in search_strings + two_cells_away_strings}
        for row in sheet.iter_rows():
            for cell in row:
                for string in search_strings:
                    if string in str(cell.value):
                        next_cell_value = sheet.cell(row=cell.row, column=cell.column + 1).value
                        if next_cell_value is not None:
                            found_values[string] = next_cell_value
                for string in two_cells_away_strings:
                    if string in str(cell.value):
                        two_cells_away_value = sheet.cell(row=cell.row, column=cell.column + 2).value
                        if two_cells_away_value is not None:
                            found_values[string] = two_cells_away_value
        return found_values

    def check_files_modified(self):
        # Check if any input files have been modified since the last output file update
        output_file_modified = False
        output_file_timestamp = 0

        if os.path.exists(self.output_file):
            output_file_timestamp = os.path.getmtime(self.output_file)

        for file_path in self.file_paths:
            file_timestamp = os.path.getmtime(file_path)
            if file_timestamp > output_file_timestamp:
                output_file_modified = True
                break

        return output_file_modified

    def combine_files(self):
        # Combine input files into a single CSV file
        output_file_modified = self.check_files_modified()
        if not output_file_modified:
            print("No changes detected in the input files. Using the existing combined file.")
            return
        # Create an empty dataframe with the desired columns
        column_names = ['Type', 'Sheet'] + search_strings + two_cells_away_strings
       
        df = pd.DataFrame(columns=column_names)

        total_sheets = sum(len(load_workbook(file_path, read_only=True).sheetnames) for file_path in self.file_paths)

        # Progress bar for file and sheet processing
        progress_bar = tqdm(total=total_sheets, desc="Processing Files and Sheets")

        for file_index, (file_path, file_type) in enumerate(zip(self.file_paths, self.file_types)):
            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheets = wb.sheetnames

            for sheet_index, sheet_name in enumerate(sheets):
                sheet = wb[sheet_name]
                found_values = self.find_values(sheet)

                temp_data = {'Type': file_type, 'Sheet': sheet_name}

                for string, value in found_values.items():
                    if string == 'Service date':
                        if value is not None:
                            if isinstance(value, datetime.datetime):
                               # value = value.date()#.isoformat()
                                value = value.date().strftime('%Y-%m-%d')
                            else:
                                try:
                                    value = datetime.datetime.strptime(str(value), '%d/%m/%Y').date().strftime('%Y-%m-%d')#.isoformat()
                                except ValueError:
                                    value = 'N/A'
                    temp_data[string] = value

                temp_df = pd.DataFrame(temp_data, index=[0])

                temp_df = temp_df[column_names]

                df = pd.concat([df, temp_df], ignore_index=True)

                progress_bar.update(1)
                time.sleep(0.1)  # Simulate processing time

        progress_bar.close()
        
        df.columns =  ['MK_Type', 'Sheet', 'Client', 'Country', 'Service_date', 'Reason_for_Service',
                            'RemScan_Serial', 'User_ID', 'User_Password','Background_Cap',
                           'Polystyrene_PS_Cap',
                            'SNR_1142_1042_cm1',
                            'SNR_2600_2500_cm1','Centre_burst_intensity',
                            'Single_beam_spectrum_4200_4500',
                            'Single_beam_spectrum_2600_3000']
        for col in ['Background_Cap',
                           'Polystyrene_PS_Cap',
                            'SNR_1142_1042_cm1',
                            'SNR_2600_2500_cm1','Centre_burst_intensity',
                            'Single_beam_spectrum_4200_4500',
                            'Single_beam_spectrum_2600_3000']:
            df[col] = df[col].apply(lambda x: x if type(x) in [int, float] else None)
        # Save the dataframe to a CSV file
        df.to_csv(self.output_file, index=False)
        


def check_valid_path(path):
    if not os.path.exists(path):
        raise ImportError('Invalid path. Check and try again.')


# Get the directory of the main Python file
directory = input('Enter a directory for the Master Instrument sheets.')
check_valid_path(directory)
print('Reading file from :', directory)

# os.path.dirname(os.path.abspath(__file__))  # Current folder

# Define the file paths and types
file_paths = [
    os.path.join(directory, 'mk1 Technical test Master copy.xlsm'),
    os.path.join(directory, 'mk2 Technical test Master copy.xlsx')
]
file_types = ['mk1', 'mk2']

# Define the specific strings you're looking for
search_strings = ['Client',
                  'Country',
                  'Service date',
                  'Reason for Service',
                  'RemScan Serial #',
                  'User ID',
                  'Password',
                  'Background Cap (Minimum requirement = 4500 @ Gain = 255)',
                  'Polystyrene P/S Cap (Minimum requirement = 4000 @ Gain = 255)',
                  'SNR: (1142 - 1042 cm-1) (Recommended requirement = 4500)',
                  'SNR: (2600 - 2500 cm-1) ',
                  'Centre burst intensity (Interferogram) (Minmum requirement =20,000)']

two_cells_away_strings = [
    'Single beam spectrum (Counts: 4200-4500 / Total Counts)x100                  (Minimum requirement =1%)',
    'Single beam spectrum (Counts: 2600-3000 / Total Counts)x100                  (Minimum requirement = 7%)']

# Construct the output file path
output_file = os.path.join(directory, 'combine.csv')

# Create an instance of the ExcelFileCombiner class
combiner = ExcelFileCombiner(file_paths, file_types, output_file, search_strings, two_cells_away_strings)


# Call the combine_files method
combiner.combine_files()
