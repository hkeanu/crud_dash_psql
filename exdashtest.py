import os
import datetime
import pandas as pd
import dash
from dash import dcc, html, Input, Output, State
from dash.exceptions import PreventUpdate
from openpyxl import load_workbook
import warnings
from tqdm import tqdm
import io
import base64
import dash_table
import numpy as np


warnings.simplefilter("ignore", category=UserWarning)

app = dash.Dash(__name__)

app.layout = html.Div([
    dcc.Upload(
        id='upload-data',
        children=html.Div([
            'Drag and Drop or ',
            html.A('Select Excel File')
        ]),
        style={
            'width': '100%',
            'height': '60px',
            'lineHeight': '60px',
            'borderWidth': '1px',
            'borderStyle': 'dashed',
            'borderRadius': '5px',
            'textAlign': 'center',
            'margin': '10px'
        },
        multiple=False
    ),
    html.Div(id='output-data-upload'),
    dcc.Loading(id="loading-output", type="default", children=[]),
    html.Div(id='table-container')
])

def process_excel(file_path):
    search_strings = [
        'Client',
        'Country',
        'Service date',
        'Reason for Service',
        'RemScan Serial #',
        'User ID',
        'Password',
        'Background Cap (Minimum requirement = 4500 @ Gain = 255)',
        'Polystyrene P/S Cap (Minimum requirement = 4000 @ Gain = 255)',
        'SNR: (1142 - 1042 cm-1) (Recommended requirement = 4500)',
        'SNR: (2600 - 2500 cm-1) ',
        'Centre burst intensity (Interferogram) (Minmum requirement =20,000)'
    ]

    two_cells_away_strings = [
        'Single beam spectrum (Counts: 4200-4500 / Total Counts)x100                  (Minimum requirement = 1%)',
        'Single beam spectrum (Counts: 2600-3000 / Total Counts)x100                  (Minimum requirement = 7%)'
    ]

    df = pd.DataFrame(columns= ['Type', 'Sheet'] +search_strings + two_cells_away_strings)

    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheets = wb.sheetnames

    for sheet_name in sheets:
        sheet = wb[sheet_name]
        found_values = find_values(sheet, search_strings, two_cells_away_strings)

        temp_data = {'Type': np.NaN, 'Sheet': np.NaN}

        for string, value in found_values.items():
            if string == 'Service date':
                if value is not None:
                    if isinstance(value, datetime.datetime):
                        value = value.date().strftime('%Y-%m-%d')
                    else:
                        try:
                            value = datetime.datetime.strptime(str(value), '%d/%m/%Y').date().strftime('%Y-%m-%d')
                        except ValueError:
                            value = 'N/A'
            temp_data[string] = value

        temp_df = pd.DataFrame(temp_data, index=[0])

        df = pd.concat([df, temp_df], ignore_index=True)

    df.columns = [
        'MK_Type', 'Sheet', 'Client', 'Country', 'Service_date', 'Reason_for_Service', 'RemScan_Serial', 'User_ID', 'User_Password',
        'Background_Cap', 'Polystyrene_PS_Cap', 'SNR_1142_1042_cm1', 'SNR_2600_2500_cm1', 'Centre_burst_intensity',
        'Single_beam_spectrum_4200_4500', 'Single_beam_spectrum_2600_3000'
    ]

    for col in [
        'Background_Cap', 'Polystyrene_PS_Cap', 'SNR_1142_1042_cm1', 'SNR_2600_2500_cm1', 'Centre_burst_intensity',
        'Single_beam_spectrum_4200_4500', 'Single_beam_spectrum_2600_3000'
    ]:
        df[col] = df[col].apply(lambda x: x if type(x) in [int, float] else None)

    return df

def find_values(sheet, search_strings, two_cells_away_strings):
    found_values = {string: None for string in search_strings + two_cells_away_strings}
    for row in sheet.iter_rows():
        for cell in row:
            for string in search_strings:
                if string in str(cell.value):
                    next_cell_value = sheet.cell(row=cell.row, column=cell.column + 1).value
                    if next_cell_value is not None:
                        found_values[string] = next_cell_value
            for string in two_cells_away_strings:
                if string in str(cell.value):
                    two_cells_away_value = sheet.cell(row=cell.row, column=cell.column + 2).value
                    if two_cells_away_value is not None:
                        found_values[string] = two_cells_away_value
    return found_values

@app.callback(
    Output('output-data-upload', 'children'),
    Output('loading-output', 'children'),
    Output('table-container', 'children'),  # New output for the table
    Input('upload-data', 'contents'),
    State('upload-data', 'filename'),
    State('upload-data', 'last_modified')
)
def update_output(contents, filename, last_modified):
    try:
        if contents is None:
            raise PreventUpdate

        directory = 'uploads'
        os.makedirs(directory, exist_ok=True)
        file_path = os.path.join(directory, filename)

        # Decode the base64 encoded content
        content_decoded = base64.b64decode(contents.split(",")[1])

        with open(file_path, 'wb') as f:
            f.write(content_decoded)

        df = process_excel(file_path)
        cleaned_csv_path = file_path.replace('.xlsx', '_cleaned.csv')
        df.to_csv(cleaned_csv_path, index=False)

        # Create the Dash table component with the cleaned CSV data
        table = dcc.Loading(id="table-loading", type="default", children=[
            dash_table.DataTable(
                id='datatable',
                columns=[{'name': col, 'id': col} for col in df.columns],
                data=df.to_dict('records'),
                page_size=10  # Adjust as needed
            )
        ])

        return [
            html.Div([
                html.H4(f'File Name: {filename}'),
                html.P('Data cleaning and extraction completed.'),
                html.A('Download Cleaned CSV', href=f'/{cleaned_csv_path}'),
            ]),
            None,
            table  # Return the table as an output
        ]
    except Exception as e:
        print(f"Error in callback: {str(e)}")
        return [
            html.Div([
                html.P('An error occurred while processing the file. Please check the file format and try again.'),
            ]),
            None,
            None  # Return None for the table in case of an error
        ]

if __name__ == '__main__':
    app.run_server(debug=True)
